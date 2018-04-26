# -*- coding: UTF-8 -*-
# with Python3
import urllib.request
import lxml.html
import openpyxl


def downloader(url):
    """下载一个网页的内容"""
    return urllib.request.urlopen(url).read()


def selector(html, regex):
    """从网页内容中筛选符合regex条件的信息，返回一类信息的列表"""
    return lxml.html.fromstring(html).cssselect(regex)


def output(*args):
    """导出获得的数据到一个xlsx文件"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '爬取表'

    value = ["url", "公司", "地方", "时间", "福利", "职位", "薪水"]

    for i in range(1, len(value)):
        sheet.cell(row=1, column=i, value=value[i - 1])

    for col, arg in enumerate(args):
        col += 1
        for row, content in enumerate(arg):
            sheet.cell(row=row + 2, column=col, value=content)

    wb.save('output/text.xlsx')
    print('数据导出成功')


def content_washer(*args):
    """数据清洗，将对象里的信息提取出来"""
    for arg in args:
        for index, item in enumerate(arg):
            arg[index] = item.text_content().strip()


def arrt_washer(**kwargs):
    """attr_washer(href=[url])"""
    for key in kwargs:                  # key实际是信息所在属性
        for arg in kwargs[key]:         # kwargs[key]实际相当于content_washer()中的args
            for index, item in enumerate(arg):
                arg[index] = item.get(key).strip()


def main():
    URL = 'http://gz.ganji.com/zpruanjiangongchengshi/zhaopin/o0/'

    url_regex = 'dt > a'                        #在job的href属性里
    company_regex = 'div.new-dl-company > a'    #在title属性里
    place_regex = 'dd.pay'
    date_regex = 'dd.pub-time > span'
    welfare_regex = 'div.new-dl-tags'
    job_regex = 'dt > a'
    salary_regex = 'dd.company > div.new-dl-salary'

    html = downloader(URL)

    url = selector(html, url_regex)         #要获取的内容与job在同一个标签内，不过链接是属性href内的信息用到的方法不一样
    company = selector(html, company_regex)
    place = selector(html, place_regex)
    date = selector(html, date_regex)
    welfare = selector(html, welfare_regex)
    job = selector(html, job_regex)
    salary = selector(html, salary_regex)

    content_washer(place, date, welfare, job, salary)
    arrt_washer(href=[url], title=[company])
    output(url, company, place, date, welfare, job, salary)


if __name__ == '__main__':
    main()
